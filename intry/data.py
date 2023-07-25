import pickle
import re
import xml.etree.ElementTree as et
from collections import defaultdict
from functools import reduce
from io import BytesIO
from itertools import groupby
from statistics import fmean
from typing import Dict, List, Optional, Set, Tuple
from urllib.request import urlopen
from zipfile import ZipFile

import openpyxl

from .intrin import *
from .tags import *
from .utils import *


class data:
    @property
    def wpkey(self) -> int:
        return self._wpkey

    @property
    def wlkey(self) -> int:
        return self._wlkey

    @property
    def wtkey(self) -> int:
        return self._wtkey

    def __init__(self, d: Dict[str, List[intrin]], wp, wl, wt):
        self._dict = d
        self.__getitem__ = self._dict.__getitem__
        self.keys = self._dict.keys
        self.values = self._dict.values
        self.items = self._dict.items
        self._wpkey = wp
        self._wlkey = wl
        self._wtkey = wt


_tses = Tuple[Set[int], int | None, str | None, tags]
_vdata, _vses = 0, 1


def get_data_src(data_source: str, *files: str) -> list[bytes]:
    if os.path.isfile(data_source):
        print_(f'importing data from {data_source}...')
        src = data_source
    elif data_source.startswith('http'):
        print_(f'downloading data from {data_source}...')
        src = BytesIO(urlopen(data_source).read())
    else:
        raise ValueError('invalid data source, expected file or url')
    with ZipFile(src) as f:
        return list(map(f.read, files))


def get_data(intel_source: Optional[str], amd_source: Optional[str]) -> data:
    path = user_data_dir('data')
    if os.path.exists(path) and not (intel_source or amd_source):
        with open(path, 'rb') as f:
            ver, dat = pickle.load(f)
            if ver == _vdata:
                return dat
    dat: Dict[str, List[intrin]] = defaultdict(list)
    fs = get_data_src(intel_source or 'https://cdrdv2.intel.com/v1/dl/getContent/671338',
                      'Intel Intrinsics Guide/files/data.js',
                      'Intel Intrinsics Guide/files/perf2.js')
    xml = fs[0].lstrip(b'var data_js = "').strip().rstrip(b'";').decode('unicode_escape')
    p2 = eval(re.sub(r'([,{])(\w+):', r'\1"\2":', fs[1].decode('U8').lstrip('perf2_js =')))
    print_('reading dat...')
    p = {}
    # p = {k: reduce(dict.__or__, chain.from_iterable(v.values()))
        #  for k, v in chain.from_iterable(map(dict.items, p_.values()))}
    for k, v in p2.items():
        p.setdefault(k.split('_')[0], {}).update(reduce(dict.__or__, v))
    ws = [0, len('Cl'), len('CPI')]
    for d in p.values():
        for k, v in d.items():
            d[k] = v.get('l', ''), v.get('t', '')
            ws = list(map(max, zip(ws, map(len, (k, *d[k])))))
    root = et.fromstring(xml)
    in2obj = defaultdict[str, list[intrin]](list)
    for i in root.findall('intrinsic'):
        def orx(x_):
            class x:
                text = None
                attrib = defaultdict(lambda: None)
            return x if x_ is None else x_
        in_ = orx(i.find('instruction')).attrib['name']
        in_ = intrin(
            i.attrib['name'],
            i.attrib['tech'],
            orx(i.find('category')).text,
            (in_ or '').lower(),
            orx(i.find('description')).text,
            [(p.attrib['type'], p.attrib['varname'])
                for p in i.findall('parameter') if p.attrib['type'] != 'void'],
            orx(i.find('return')).attrib['type'],
            orx(i.find('header')).text,
            (orx(i.find('operation')).text or '').strip(),
            p.get((in_ or '').upper(), {}))
        in2obj[in_.instr].append(in_)
        dat[i.attrib['tech']].append(in_)
    dat = dict(sorted((k, sorted(v)) for k, v in dat.items()))
        
    fs = get_data_src(amd_source or 'https://www.amd.com/system/files/TechDocs/57647.zip', 'Zen4_Instruction_Latencies_version_1-00.xlsx')
    wb = openpyxl.open(BytesIO(fs[0]), read_only=True)
    rs = iter(wb.get_sheet_by_name('Zen4 instruction latencies'))
    ii, il, it = map([c.value for c in next(rs)].index, ('Instruction', 'Latency', 'Throughput'))
    def gavg(xs: Iterable[str | float | int]) -> float:
        def map_(x: str | float | int) -> float | None:
            with suppress(ValueError):
                return float(x)
            return (m := re.match(r'([\d.]+)-([\d.]+)', str(x))) and (float(m[1]) + float(m[2])) / 2
        return fmean(xs) if (xs := list(filter(None, map(map_, xs)))) else 0
    newws = ws[1:]
    for k, g in groupby(rs, lambda r: r[ii].value.lower()):
        g = list(g)
        l = set(str(r[il].value) for r in g if r[il].value)
        t = set(str(r[it].value) for r in g if r[it].value)
        ls = ls if len(ls := '/'.join(l)) <= ws[1] else f'~{gavg(l):.2f}'
        ts = ts if len(ts := '/'.join(t)) <= ws[2] else f'~{gavg(t):.2f}'
        for in_ in in2obj.get(k, []):
            in_.perf['Zen4'] = (ls, ts)
        newws = list(map(max, zip(newws, map(len, (ls, ts)))))
    ws[1:] = newws
    print_('writing to disk...')
    dat = data(dat, ws[0], ws[1], ws[2])
    with open(path, 'wb') as f:
        pickle.dump((_vdata, dat), f)
    return dat


def get_ses() -> _tses:
    path = user_data_dir('ses')
    if os.path.isfile(path):
        with open(path, 'rb') as f:
            ver, ses = pickle.load(f)
            if ver == _vses:
                return ses
    return set(), None, None, tags()


def dump_ses(ses: _tses):
    path = user_data_dir('ses')
    with open(path, 'wb') as f:
        pickle.dump((_vses, ses), f)
