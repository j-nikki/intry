from argparse import ArgumentParser
from contextlib import suppress
import json
import os
import pickle
import re
from sqlite3 import connect
import xml.etree.ElementTree as et
from collections import defaultdict
from functools import reduce
from io import BytesIO
from itertools import groupby
from statistics import fmean
from typing import Dict, Iterable, List, Optional
from urllib.request import urlopen
from zipfile import ZipFile

import openpyxl


class intrin:
    def __init__(self, name, tech, cat, instr, descr, params, ret, header, op, perf):
        self._name = name
        self._tech = tech
        self._cat = cat
        self._instr = instr
        self._descr = descr
        self._params = params
        self._ret = ret
        self._header = header
        self._op = op
        self._perf = perf

    @property
    def name(self) -> str:
        return self._name

    @property
    def tech(self) -> str:
        return self._tech

    @property
    def cat(self) -> str:
        return self._cat

    @property
    def instr(self) -> str:
        return self._instr

    @property
    def descr(self) -> str:
        return self._descr

    @property
    def params(self) -> list:
        return self._params

    @property
    def ret(self) -> str:
        return self._ret

    @property
    def header(self) -> str:
        return self._header

    @property
    def op(self) -> str:
        return self._op

    @property
    def perf(self) -> dict[str, tuple[str, str]]:
        """mapping of microarchitecture to <latency, throughput> pair"""
        return self._perf

    def __lt__(self, other):
        return self.name < other.name


class data:
    @property
    def wpkey(self) -> int:
        return self._wpkey

    @property
    def wlkey(self) -> int:
        return self._wlkey

    @property
    def wtkey(self) -> int:
        return self._wtkey

    def __init__(self, d: Dict[str, List[intrin]], wp, wl, wt):
        self._dict = d
        self.__getitem__ = self._dict.__getitem__
        self.keys = self._dict.keys
        self.values = self._dict.values
        self.items = self._dict.items
        self._wpkey = wp
        self._wlkey = wl
        self._wtkey = wt


DATA_VER = 0


def _fetch_data(data_source: str, *files: str) -> list[bytes]:
    if os.path.isfile(data_source):
        print(f"importing data from {data_source}...")
        src = data_source
    elif data_source.startswith("http"):
        print(f"downloading data from {data_source}...")
        src = BytesIO(urlopen(data_source).read())
    else:
        raise ValueError("invalid data source, expected file or url")
    with ZipFile(src) as f:
        return list(map(f.read, files))


def get_data(path, intel_source: Optional[str], amd_source: Optional[str]) -> data:
    if os.path.exists(path) and not (intel_source or amd_source):
        with open(path, "rb") as f:
            ver, dat = pickle.load(f)
            if ver == DATA_VER:
                return dat
    dat: Dict[str, List[intrin]] = defaultdict(list)
    fs = _fetch_data(
        intel_source or "https://cdrdv2.intel.com/v1/dl/getContent/671338",
        "Intel Intrinsics Guide/files/data.js",
        "Intel Intrinsics Guide/files/perf2.js",
        "Intel Intrinsics Guide/files/notes.json",
    )
    xml = (
        fs[0].lstrip(b'var data_js = "').strip().rstrip(b'";').decode("unicode_escape")
    )
    p2 = eval(
        re.sub(r"([,{])(\w+):", r'\1"\2":', fs[1].decode("U8").lstrip("perf2_js ="))
    )
    notes = {
        n["note_name"]: re.sub(
            r"([\s])\1+", r"\1", re.sub(r"<[^<>]+?>", "\n", n["note_value"])
        ).strip()
        for n in json.loads(fs[2].decode("U8"))["notes"]
    }
    pnote = re.compile(rf'^\s*\[({"|".join(notes.keys())})\]', flags=re.M)
    print("reading dat...")
    p = {}
    # p = {k: reduce(dict.__or__, chain.from_iterable(v.values()))
    #  for k, v in chain.from_iterable(map(dict.items, p_.values()))}
    for k, v in p2.items():
        p.setdefault(k.split("_")[0], {}).update(reduce(dict.__or__, v))
    ws = [0, len("Cl"), len("CPI")]
    for d in p.values():
        for k, v in d.items():
            d[k] = v.get("l", ""), v.get("t", "")
            ws = list(map(max, zip(ws, map(len, (k, *d[k])))))
    root = et.fromstring(xml)
    in2obj = defaultdict[str, list[intrin]](list)
    for i in root.findall("intrinsic"):

        def orx(x_):
            class x:
                text = None
                attrib = defaultdict(lambda: None)

            return x if x_ is None else x_

        in_ = orx(i.find("instruction")).attrib["name"]
        in_ = intrin(
            i.attrib["name"],
            i.attrib["tech"],
            orx(i.find("category")).text,
            (in_ or "").lower(),
            pnote.sub(lambda m: notes[m[1]], orx(i.find("description")).text),
            [
                (p.attrib["type"], p.attrib["varname"])
                for p in i.findall("parameter")
                if p.attrib["type"] != "void"
            ],
            orx(i.find("return")).attrib["type"],
            orx(i.find("header")).text,
            (orx(i.find("operation")).text or "").strip(),
            p.get((in_ or "").upper(), {}),
        )
        in2obj[in_.instr].append(in_)
        dat[i.attrib["tech"]].append(in_)
    dat = dict(sorted((k, sorted(v)) for k, v in dat.items()))

    fs = _fetch_data(
        amd_source or "https://www.amd.com/system/files/TechDocs/57647.zip",
        "Zen4_Instruction_Latencies_version_1-00.xlsx",
    )
    wb = openpyxl.open(BytesIO(fs[0]), read_only=True)
    rs = iter(wb.get_sheet_by_name("Zen4 instruction latencies"))
    ii, il, it = map(
        [c.value for c in next(rs)].index, ("Instruction", "Latency", "Throughput")
    )
    prng = re.compile(r"([\d.]+)-([\d.]+)")

    def gavg(xs: Iterable[str | float | int]) -> float:
        def map_(x: str | float | int) -> float | None:
            with suppress(ValueError):
                return float(x)
            return (m := prng.match(str(x))) and (float(m[1]) + float(m[2])) / 2

        return fmean(xs) if (xs := list(filter(None, map(map_, xs)))) else 0

    newws = ws[1:]
    for k, g in groupby(rs, lambda r: r[ii].value.lower()):
        g = list(g)
        l = set(str(r[il].value) for r in g if r[il].value)
        t = set(str(r[it].value) for r in g if r[it].value)
        ls = ls if len(ls := "/".join(l)) <= ws[1] else f"~{gavg(l):.2f}"
        ts = ts if len(ts := "/".join(t)) <= ws[2] else f"~{gavg(t):.2f}"
        for in_ in in2obj.get(k, []):
            in_.perf["Zen4"] = (ls, ts)
        newws = list(map(max, zip(newws, map(len, (ls, ts)))))
    ws[1:] = newws
    print("writing to disk...")
    dat = data(dat, ws[0], ws[1], ws[2])
    with open(path, "wb") as f:
        pickle.dump((DATA_VER, dat), f)
    return dat


schema = """
PRAGMA foreign_keys = ON;
PRAGMA journal_mode = WAL;
PRAGMA synchronous = NORMAL;

BEGIN;

CREATE TABLE IF NOT EXISTS intrin (
    id INTEGER PRIMARY KEY,
    name TEXT NOT NULL UNIQUE,
    tech TEXT NOT NULL,
    cat TEXT NOT NULL,
    instr TEXT NOT NULL,
    descr TEXT NOT NULL,
    ret TEXT NOT NULL,
    header TEXT NOT NULL,
    op TEXT NOT NULL
) STRICT;

CREATE TABLE IF NOT EXISTS param (
    intrin_id INTEGER NOT NULL REFERENCES intrin(id) ON DELETE CASCADE,
    number INTEGER NOT NULL,
    type TEXT NOT NULL,
    name TEXT NOT NULL,
    PRIMARY KEY (intrin_id, number)
) STRICT;

CREATE TABLE IF NOT EXISTS perf (
    intrin_id INTEGER NOT NULL REFERENCES intrin(id) ON DELETE CASCADE,
    arch TEXT NOT NULL,
    latency TEXT NOT NULL,
    throughput TEXT NOT NULL,
    PRIMARY KEY (intrin_id, arch)
) STRICT;

COMMIT;
"""

if __name__ == "__main__":
    ap = ArgumentParser()
    ap.add_argument("--intel-source", "-i", help="intel data source")
    ap.add_argument("--amd-source", "-a", help="amd data source")
    args = ap.parse_args()
    datadir = (
        os.getenv("XDG_DATA_HOME", os.path.expanduser("~/.local/share")) + "/intry"
    )
    os.makedirs(datadir, exist_ok=True)
    db = datadir + "/db"
    ws = datadir + "/ws"
    with connect(db) as conn:
        conn.executescript(schema)
        c = conn.cursor()
        dat = get_data(datadir + "/data", args.intel_source, args.amd_source)
        print("inserting data...")
        for tech, intrins in dat.items():
            for in_ in intrins:
                (iid,) = c.execute(
                    "INSERT OR REPLACE INTO intrin (name, tech, cat, instr, descr, ret, header, op) VALUES (?, ?, ?, ?, ?, ?, ?, ?) RETURNING id",
                    (
                        in_.name,
                        tech,
                        in_.cat,
                        in_.instr,
                        in_.descr.replace("\t", "    "),
                        in_.ret,
                        in_.header,
                        in_.op.replace("\t", "    "),
                    ),
                ).fetchone()
                c.executemany(
                    "INSERT OR REPLACE INTO param VALUES (?, ?, ?, ?)",
                    ((iid, i, *p) for i, p in enumerate(in_.params)),
                )
                c.executemany(
                    "INSERT OR REPLACE INTO perf VALUES (?, ?, ?, ?)",
                    ((iid, a, l, t) for a, (l, t) in in_.perf.items()),
                )
        with open(ws, "wb") as f:
            data = b"".join(map(int.to_bytes, (dat.wpkey, dat.wlkey, dat.wtkey)))
            f.write(data)
